import requests
from PIL import Image
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from io import BytesIO
import openpyxl
import os

path1 = 'C:\\Users\\Asus\\Desktop\\XD\\python\\Pokeapi\\requests-api\\Output.txt'
legendario = {'true': 0, 'false': 0}
tipopokemon = {}
gen = {}

def main():
    nompokemon = input("Bienvenido a la Pokédex, inserte el nombre del Pokémon: ")
    
    with open(path1, "a") as text_file:
        print(f"{nompokemon}", file=text_file)
    
    pokemon = busquedapokemon(nompokemon)
    otropoke = species(nompokemon)

    if pokemon:
        while True:
            try:

                menu()
                opc = int(input("Escoja una opción: "))

            except ValueError:

                print("Por favor escriba una opción válida.")
                continue

            else:
                if opc == 10:
                    print("Saliendo del programa...")
                    break
                
                elif opc < 1 or opc > 10:
                    print("Opción inválida. Por favor, seleccione una opción válida.")
                    continue

                elif opc == 7:
                    with open(path1, "w") as text_file:
                        print("", file=text_file)
                
                else:
                    mostrar_informacion(opc, pokemon, otropoke)


def menu():
    print("\n1. Peso\n2. Altura\n3. Tipo\n4. Imagen\n5. ¿Es legendario?\n6. Gráficas\n7. Borrar requests\n8. Generación\n9. Crear excel de datos\n10. Cerrar programa")


def busquedapokemon(busqueda):
    url = f"https://pokeapi.co/api/v2/pokemon/{busqueda.lower()}"
    response = requests.get(url)

    if response.status_code == 200:
        return response.json()
    else:
        print("No se encontró el Pokémon buscado.")
        return None
    
def species(busq):
    url = f"https://pokeapi.co/api/v2/pokemon-species/{busq.lower()}"
    response = requests.get(url)
    if response.status_code == 200:
        return response.json()
    
def mostrar_informacion(opcion, datos_pokemon, spec):

    if opcion == 1:
        print(f"Peso: {datos_pokemon['weight']} kg")
        with open(path1, "a") as text_file:
            print(f"Peso: {datos_pokemon['weight']}", file=text_file)

    elif opcion == 2:
        print(f"Altura: {datos_pokemon['height']} cm")

        with open(path1, "a") as text_file:
            print(f"Altura: {datos_pokemon['height']}", file=text_file)

    elif opcion == 3:
        tipos = [tipo['type']['name'] for tipo in datos_pokemon['types']]
        print("Tipos:")
        
        for tipoz in tipos:
            print(f"- {tipoz}")
            
            if tipoz not in tipopokemon:
                tipopokemon[tipoz] = 1
            else:
                tipopokemon[tipoz] += 1

        with open(path1, "a") as text_file:
            for tipo in tipos:
                print(f"- {tipo}", file=text_file)

    elif opcion == 4:
        mostrar_imagen(datos_pokemon)

    elif opcion == 5:
        print(f"{spec['is_legendary']}")

        with open(path1, "a") as text_file:
           print(f"¿Es legendario?: {spec['is_legendary']}", file=text_file)

        if spec['is_legendary'] == True:
            legendario['true'] += 1
        if spec['is_legendary'] == False:
            legendario['false'] += 1

    elif opcion == 6:
        graficas()
    
    elif opcion == 8:
        generacion = spec['generation']['name']
        print(f"{generacion}")

        with open(path1, "a") as text_file:
           print(f"Generación: {generacion}", file=text_file) 
        
        if generacion not in gen:
            gen[generacion] = 1
        else:
            gen[generacion] += 1

    elif opcion == 9:
        print("Creando el excel...")
        pyxl()

def mostrar_imagen(datos_pokemon):
    imagen_url = datos_pokemon['sprites']['front_default']
    imagen_respuesta = requests.get(imagen_url)
    if imagen_respuesta.status_code == 200:
        imagen = Image.open(BytesIO(imagen_respuesta.content))
        plt.title(datos_pokemon['name'])
        imgplot = plt.imshow(imagen)
        plt.show()
    else:
        print("No se pudo obtener la imagen del Pokémon")

def graficas():
    #obtenemos nuestra matriz para centrar cada figura
    fig = plt.figure(figsize=(10, 10))
    gs = gridspec.GridSpec(4,4)

    #figura 1
    nombres = list(legendario.keys())
    valores = list(legendario.values())
    ax1 = plt.subplot(gs[:2, :2])
    ax1.set_title('Pokémon legendarios')
    ax1.bar(nombres, valores)

    #figura 2
    nom2 = list(tipopokemon.keys())
    val2 = list(tipopokemon.values())
    ax2 = plt.subplot(gs[:2, 2:])
    ax2.set_title('Tipos de pokémon registrados')
    ax2.bar(nom2, val2)
    
    #figura 3
    nom3 = list(gen.keys())
    val3 = list(gen.values())
    ax3 = plt.subplot(gs[2:4, 1:3])
    ax3.set_title('Generación de los pokémon')
    ax3.pie(val3, labels=nom3)

    plt.show()

def pyxl():
    if not os.path.isfile('datosnum.xlsx'):
        wb = openpyxl.Workbook()  
        dest_filename = 'datosnum.xlsx' 
        wb.save(os.path.join("C:\\Users\\Asus\\Desktop\\XD\\python\\Pokeapi\\datosnumericos-excel", dest_filename))

    sheet = wb.active
    sheet.title = 'Legendarios'
    # pokemon legendarios
    leg = list(map(list, legendario.items()))
    for dato in leg:
        sheet.append(dato)

    # tipos de pokemon
    sheet2 = wb.create_sheet('Tipos')
    tip = list(map(list, tipopokemon.items()))
    for dato2 in tip:
        sheet2.append(dato2)

    sheet3 = wb.create_sheet('Generaciones')
    generacion = list(map(list, gen.items()))
    for dato3 in generacion:
        sheet3.append(dato3)

    
    wb.save("C:\\Users\\Asus\\Desktop\\XD\\python\\Pokeapi\\datosnumericos-excel\\datosnum.xlsx")
